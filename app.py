import streamlit as st
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

# ========= CONFIG: HEADER COLOURS (MATCHING YOUR SCREENSHOT) =========
# A: green, B: dark blue, C: yellow, D: light blue, F: red, H: orange, I: grey
# Others + the rest stay white (no fill)
SHELF_COLORS = {
    "A": "#00B050",  # green
    "B": "#0070C0",  # blue
    "C": "#FFFF00",  # yellow
    "D": "#61CBF3",  # light blue
    "E": None,
    "F": "#CC0000",  # red
    "G": None,
    "H": "#FFC000",  # orange
    "I": "#808080",  # grey
    "J": None,
    "K": None,
    "L": None,
    "M": None,
    "N": None,
    "O": None,
    "Others": None,  # header white; row colours per-label from source
}

SHELF_ORDER = list("ABCDEFGHIJKLMNO") + ["Others"]


# ========= HELPERS =========

def detect_shelf(label: str) -> str:
    """
    Shelf is digit number 9 in the label ID.
    Example: HAZ-A101I123 -> 9th char = 'I' -> shelf I.
    Anything not A–O goes to Others.
    """
    if label is None:
        return "Others"

    text = str(label).strip()
    if len(text) < 9:
        return "Others"

    ch = text[8].upper()  # 0-based index → 9th character

    if ch in SHELF_ORDER[:-1]:  # A–O
        return ch
    return "Others"


def get_cell_color_hex(cell):
    """
    Get cell background colour as #RRGGBB.

    - If we have an explicit RGB (most manual fills), use it.
    - If it's theme/indexed but has a visible fill, we fall back
      to a default baby blue (#61CBF3) instead of white.
    """
    fill = cell.fill
    if not fill or getattr(fill, "patternType", None) in (None, "none"):
        return None

    # Prefer fgColor, fall back to start_color
    color_obj = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
    if not color_obj:
        return None

    ctype = getattr(color_obj, "type", None)
    rgb = getattr(color_obj, "rgb", None)

    # Direct RGB (e.g. "FFRRGGBB")
    if ctype == "rgb" and rgb:
        if len(rgb) == 8:  # ARGB
            rgb = rgb[2:]
        if len(rgb) == 6:
            return "#" + rgb.upper()

    # Theme/indexed colours – treat as baby blue so they aren't white
    if ctype in ("theme", "indexed"):
        return "#61CBF3"  # fallback baby blue

    return None


def process_sheet(ws):
    """
    Read a single worksheet.
    - All non-empty cells are labels.
    - Exclude any label that contains 'bin' (case-insensitive).
    - Shelves A–O: just text.
    - Others: keep text + source cell colour.

    Returns:
      df_out       : DataFrame with columns A–O + Others
      others_colors: list of colours for the Others column (same length as column)
    """
    groups = {shelf: [] for shelf in SHELF_ORDER}
    others_colors = []

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is None:
                continue

            text = str(val).strip()
            if not text:
                continue

            # Exclude labels that include the word "bin" (any case)
            if "bin" in text.lower():
                continue

            shelf = detect_shelf(text)
            if shelf == "Others":
                groups["Others"].append(text)
                others_colors.append(get_cell_color_hex(cell))
            else:
                groups[shelf].append(text)

    # Determine max column length
    max_len = max((len(v) for v in groups.values()), default=0)

    data = {}
    padded_others_colors = list(others_colors)
    for shelf in SHELF_ORDER:
        col_vals = groups[shelf]
        pad = max_len - len(col_vals)
        if pad > 0:
            col_vals = col_vals + [""] * pad
            if shelf == "Others":
                padded_others_colors = padded_others_colors + [None] * pad
        data[shelf] = col_vals

    df_out = pd.DataFrame(data)
    return df_out, padded_others_colors


def write_output_workbook(sheets_data):
    """
    sheets_data: { sheet_name -> (df_out, others_colors) }

    For each sheet:
      - Row 1: colour hex (if defined in SHELF_COLORS) + same bg colour.
      - Row 2: shelf letter with same bg colour.
      - Row 3+: labels.
      - Others column cells get their original colours (per label).
    """
    from pandas import ExcelWriter

    output = BytesIO()
    with ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book

        others_col_idx = SHELF_ORDER.index("Others")

        for sheet_name, (df_out, others_colors) in sheets_data.items():
            # Write data WITHOUT headers, starting row 2 (Excel row 3)
            df_out.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
                startrow=2,
            )

            ws = writer.sheets[sheet_name]

            # Header rows
            for col_idx, shelf in enumerate(SHELF_ORDER):
                header_color = SHELF_COLORS.get(shelf)

                # Row 1: colour hex
                row1_fmt = workbook.add_format(
                    {"align": "center", "valign": "vcenter", "border": 1}
                )
                # Row 2: shelf letter
                row2_fmt = workbook.add_format(
                    {"bold": True, "align": "center", "valign": "vcenter", "border": 1}
                )

                if header_color:
                    row1_fmt.set_bg_color(header_color)
                    row2_fmt.set_bg_color(header_color)
                    ws.write(0, col_idx, header_color, row1_fmt)
                else:
                    ws.write(0, col_idx, "", row1_fmt)

                ws.write(1, col_idx, shelf, row2_fmt)

            # Apply original colours to Others column cells (row 3+)
            for i, color_hex in enumerate(others_colors):
                if not color_hex:
                    continue
                value = df_out.iloc[i, others_col_idx]
                cell_fmt = workbook.add_format({"border": 1})
                cell_fmt.set_bg_color(color_hex)
                ws.write(2 + i, others_col_idx, value, cell_fmt)

            # Nice column width
            ws.set_column(0, len(SHELF_ORDER) - 1, 22)

    output.seek(0)
    return output.getvalue()


# ========= STREAMLIT APP =========

st.title("Shelf Layout Generator (Exact Header Colours + Others Preserved)")

st.write(
    """
Upload an **Excel file (.xlsx)** with bin label IDs.

For **each sheet** in the file:
- All non-empty cells are treated as label IDs.
- Any label containing the word **"bin"** (any case) is **excluded**.
- Shelf is from **digit 9** of the label.
- Labels go into columns **A–O** or **Others**.
- Header colours for A–I match your template.
- **Others**:
  - Header stays white.
  - Each label keeps its original cell colour from the uploaded sheet.
"""
)

uploaded_file = st.file_uploader(
    "Upload Excel file with bin labels (all sheets will be processed)",
    type=["xlsx"],
)

generate = st.button("Generate shelf layout Excel")

if generate:
    if uploaded_file is None:
        st.error("Please upload an Excel (.xlsx) file first.")
    else:
        try:
            wb = load_workbook(uploaded_file, data_only=True)
        except Exception as e:
            st.error(f"Failed to read Excel file: {e}")
        else:
            sheets_data = {}
            for ws in wb.worksheets:
                df_out, others_colors = process_sheet(ws)
                sheets_data[ws.title] = (df_out, others_colors)

            if not sheets_data:
                st.error("No data found in the uploaded workbook.")
            else:
                # Preview first sheet
                first_name = next(iter(sheets_data.keys()))
                preview_df, _ = sheets_data[first_name]
                st.write(f"Preview from sheet: **{first_name}**")
                st.dataframe(preview_df.head(20))

                excel_bytes = write_output_workbook(sheets_data)
                st.success("Shelf layout Excel generated successfully.")

                st.download_button(
                    label="Download shelf layout Excel",
                    data=excel_bytes,
                    file_name="shelf_labels_layout_colours.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
